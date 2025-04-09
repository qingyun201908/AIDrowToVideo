import os
import glob
import datetime
from openai import OpenAI
import openpyxl

# 初始化OpenAI客户端
client = OpenAI(
    api_key="",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# 初始化Excel文件
excel_file = "D://text_processing_log.xlsx"
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
else:
    wb = openpyxl.Workbook()
ws = wb.active
ws.title = "处理记录"
if ws.max_row == 1:
    ws.append(["类型", "内容", "文件路径", "处理时间"])

def save_to_excel(question, answer, file_path):
    """保存到Excel文件"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(["问题", question, file_path, timestamp])
    ws.append(["回答", answer, file_path, timestamp])
    wb.save(excel_file)

def print_dialog(question, answer, max_width=80):
    """在CMD界面美观打印对话"""
    sep_line = '─' * max_width
    print(f"\n{sep_line}")
    print(" 问题 ".center(max_width, '░'))
    print(f"{question}\n")
    print(" 回答 ".center(max_width, '░'))
    print(f"{answer}")
    print(sep_line)

def log_error(log_file, message, error=None):
    """统一记录错误日志"""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_file, 'a', encoding='utf-8') as f:
        log_entry = f"[{timestamp}] {message}"
        if error:
            log_entry += f"\n错误详情: {str(error)}"
        log_entry += "\n" + "-"*50 + "\n"
        f.write(log_entry)

def process_file(file_path, log_file):
    """处理单个文件"""
    try:
        if os.path.getsize(file_path) == 0:
            print(f"跳过空文件: {file_path}")
            return

        with open(file_path, 'r+', encoding='utf-8') as f:
            original_lines = f.readlines()
            
            if not any(line.strip() for line in original_lines):
                print(f"跳过内容为空的文件: {file_path}")
                return

            f.seek(0)
            f.truncate()
            
            for line_num, line in enumerate(original_lines, 1):
                try:
                    if not line.strip():
                        f.write(line)
                        continue
                    
                    original_text = line.strip()
                    
                    completion = client.chat.completions.create(
                        model="qwen-plus",
                        messages=[{
                            'role': 'user',
                            'content': f'''原本句子中存在错词，不常见词语。可以修改词语和文字不采用偏僻词，不涵盖英语词汇，请严格按正常语序重组句子，只返回修改后的句子，不要任何解释。需要修改的句子：
{original_text}'''
                        }]
                    )
                    
                    modified = completion.choices[0].message.content
                    f.write(modified + line[len(line.rstrip('\n\r')):])
                    
                    # 显示对话并保存
                    print_dialog(original_text, modified)
                    save_to_excel(original_text, modified, file_path)
                    
                except Exception as e:
                    error_msg = f"文件: {file_path} 第{line_num}行处理失败 - 保留原内容"
                    print(f"\n! 处理异常: {error_msg}")
                    log_error(log_file, error_msg, e)
                    f.write(line)

    except Exception as e:
        error_msg = f"文件处理失败: {file_path}"
        print(f"\n! 严重错误: {error_msg}")
        log_error(log_file, error_msg, e)

def process_folder(folder_path):
    """处理整个文件夹"""
    log_file = os.path.join(folder_path, "processing_errors.log")
    
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("文本处理异常日志\n")
        f.write(f"开始时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*60 + "\n")
    
    txt_files = glob.glob(os.path.join(folder_path, '**', '*.txt'), recursive=True)
    
    if not txt_files:
        print(f"在 {folder_path} 中未找到txt文件")
        return
    
    print(f"开始批量处理，共发现 {len(txt_files)} 个文本文件")
    print(f"处理记录将保存至: {os.path.abspath(excel_file)}")
    print("="*60)
    
    for i, file_path in enumerate(txt_files, 1):
        print(f"\n正在处理文件 ({i}/{len(txt_files)})：{os.path.basename(file_path)}")
        process_file(file_path, log_file)
    
    print("\n" + "="*60)
    print(f"处理完成！错误日志已保存至: {log_file}")
    print(f"共处理 {len(txt_files)*2} 条记录到Excel文件")

if __name__ == "__main__":
    try:
        process_folder('D:\\BaiduNetdiskDownload\\001-050_images')
    finally:
        wb.close()